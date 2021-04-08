from flask import Flask, jsonify, request, json, render_template
import sys
from openpyxl import load_workbook
import excel_db

app = Flask(__name__)

# 챗봇 사용자정보 관리 파일
EXCEL_FILE_NAME = 'database.xlsx'

db = load_workbook(filename=EXCEL_FILE_NAME)
user_db = db['User']
lecture_db = db['Lecture']


@app.route("/")
def hello():
    #return "Hello, macgyver on goorm.io!"
    return render_template("hello.html")

@app.route("/listlevel", methods=["get"])
def listLevel():
    userLevel = request.args.get("username")
    print(userLevel)
    
    #return render_template("hello.html")


@app.route("/searchlevel", methods = ["get"])
def searchLevel():
    
    # 추후 content는 카카오톡 리턴값으로 치환예정 @ 2021.02.11.
    content = ""
    usernm = request.args.get("username")
    
    print('searchLevel started... --> ' +usernm)
    
    # 엑셀로 사용자 정보 관리
    for idx, row in enumerate(user_db.rows):
        
        print('[' + str(idx) + ' / ' + str(user_db.max_row) + '] ' + str(row[0].value) + ' , ' + str(row[1].value) + ' , ' + str(row[2].value))
        
        if idx != 0 and row[2].value == usernm:
            user_row = row
            print('[1] user_row catched')
            break
      
        elif idx ==user_db.max_row - 1:
            NEW_INDEX = user_db.max_row + 1
            user_db[NEW_INDEX][0].value = '테스트_유져키'
            user_db[NEW_INDEX][1].value = 0
            db.save(EXCEL_FILE_NAME)
            user_row = user_db[user_db.max_row]
            
            print('[2] user_row catched and db saved')
            
            response = {
                "message": {
                    "text": "{name}님, 처음 방문 감사합니다. 본인의 직군/직종/직무를 입력해주시겠어요?".format(name=usernm)
                },
                "keyboard": {
                    "type": "text"
                }
            }
            return jsonify(response)
    
    if user_row[1].value is 0 :
        user_row[1].value = 1
        user_row[2].value = usernm
        db.save(EXCEL_FILE_NAME)
        
    tags = request.args.get("hashtag")
    
    print("HR hashtag --> " + tags)
    
    # 엑셀로 강좌(콘텐츠목록) 관리 @ 2021.02.09.
    for idx, row in enumerate(lecture_db.rows):
        
        print('[' + str(idx) + ' / ' + str(lecture_db.max_row) + '] ' + str(row[0].value) + ' , ' + str(row[1].value) + ' , ' + str(row[2].value) + str(row[3].value) + str(row[4].value))
        
        if idx != 0 and tags in row[0].value:
            lecture_row = row
            print('[1] lecture_row catched')
            break    
        else:
            lecture_row = 0
            continue
        
    # excel_db에서 가져온 강의(콘텐츠) 목록이 1건이상 있으면, response 생성
    if lecture_row != 0:    

        response = {
            "message": {
                "text" : "{name}님, 또 오셨군요 ㅋㅋ \n www.youtube.com/embed/{video_link} 추천! \n 제목 : {title} \n 채널명 : {channelnm}".replace("\n", "\\r\\n").format(name=usernm, video_link=lecture_row[1].value, title=lecture_row[4].value, channelnm=lecture_row[3].value)
            },
            "keyboard": {
                "type": "buttons",
                "buttons": ["커뮤니티소개", "콘텐츠표", "홈으로"]
            }
        }
        return jsonify(response)
    
    #엑셀로 카카오톡 기본 UI 구현 @ 2021.02.06.
    #콘텐츠 추천부분 추가 @ 2021.02.07.
    try :
        # 콘텐츠 추천
        if content == u"콘텐츠소개":
            if user_row[3].value is not None:
                level = user_row[3].value
                response = excel_db.get_lectures(level, user_row)
            else :
                response = {
                    "message" : {
                        "text":"학습 수준을 알려주세요."
                    },
                    "keyboard" : {
                        "type": "buttons",
                        "buttons":["초급", "중급","고급"]
                    }
                }
        elif content in ["초급", "중급", "고급"]:
            user_row[3].value = content
            db.save(EXCEL_FILE_NAME)
            response = excel_db.get_lectures(content, user_row)
        else:
            response = excel_db.get_lectures(content, user_row)
    except:
        response = {
            "message": {
                "text" : "다시 시도해 주세요."
            },
            "keyboard": {
                "type": "buttons",
                "buttons": ["홈으로"]
            }
        }
        
    return jsonify(response)
    

# ======= 여기서부턴 카카오톡 챗봇 연동 메소드 정의부 =========


@app.route("/keyboard")
def keyboard():
    
    response = {
        "type": "button",
            "buttons": ["홈으로"]
    }
    return jsonify(response)



@app.route("/message", methods = ["POST"])
def message():
    
    data = json.loads(request.data)
    
    content = data["content"]
    user_key = data["user_key"]
    
    # 엑셀로 사용자 정보 관리
    for idx, row in enumerate(user_db.rows):
        if idx != 0 and row[0].value == user_key:
            user_row = row
            break
      
        if idx ==user_db.max_row - 1:
            NEW_INDEX = user_db.max_row + 1
            user_db[NEW_INDEX][0].value = user_key
            user_db[NEW_INDEX][1].value = 0
            db.save(EXCEL_FILE_NAME)
            user_row = user_db[user_db.max_row]
            
            response = {
                "message": {
                    "text": "처음 방문하셨네요. 이름이 (직군이, 직종이, 직무가, 그 외 관심사가..) 어떻게 되세요?"
                },
                "keyboard": {
                    "type": "text"
                }
            }
            return jsonify(response)
            
        if user_row[1].value is 0 :
            user_row[1].value = 1
            user_row[2].value = content
            db.save(EXCEL_FILE_NAME)
            
            response = {
                "message": {
                    "text" : "{name}님, 반갑습니다.".format(name=content)
                },
                "keyboard": {
                    "type": "buttons",
                    "buttons": ["커뮤니티소개", "콘텐츠소개", "홈으로"]
                }
            }
            return jsonify(response)
            
    #카카오톡 기본 UI 코딩 --> 엑셀에 콘텐츠 넣어 연동(excel_db) 전환에 따른 주석 @ 2021.02.06.
    '''
    if content == u"홈으로":
        response = {
            "message" : {
                "text" : "원하시는 정보 버튼 눌러주세요."
            },
            "keyboard": {
                "type": "buttons",
                "buttons": ["커뮤니티소개", "콘텐츠표","홈으로"]
            }
        }
    elif content == u"커뮤니티소개":
        response = {
            "message": {
                "text": "HIS 유관 생태계 협업 플랫폼 !"
            },
            "keyboard": {
                "type" : "buttons",
                "buttons": ["홈으로"]
            }
        }
    elif content == u"콘텐츠표":
        response = {
            "message": {
                "text": "다샤 콘텐츠 목록입니다.",
                "message_button": {
                    "label": "웹링크로 콘텐츠목록 보기",
                    "url": "https://www/ai-academy.ai/ai-b2c"
                }
            },
            "keyboard": {
                "type": "buttons",
                "buttons":["홈으로"]
            }
        }
    '''
        
    '''
    response = {
        "message": {
            "text": "Hello, World"
        }
    }
    '''
    
    #엑셀로 카카오톡 기본 UI 구현 @ 2021.02.06.
    #콘텐츠 추천부분 추가 @ 2021.02.07.
    try :
        # 콘텐츠 추천
        if content == u"콘텐츠소개":
            if user_row[3].value is not None:
                level = user_row[3].value
                response = excel_db.get_lectures(level, user_row)
            else :
                response = {
                    "message" : {
                        "text":"학습 수준을 알려주세요."
                    },
                    "keyboard" : {
                        "type": "buttons",
                        "buttons": ["초급", "중급", "고급"]
                    }
                }
        elif content in ["초급", "중급", "고급"]:
            user_row[3].value = content
            db.save(EXCEL_FILE_NAME)
            response = excel_db.get_lectures(content, user_row)
        else:
            response = excel_db.get_lectures(content, user_row)
    except:
        response = {
            "message": {
                "text": "다시 시도해 주세요."
            },
            "keyboard": {
                "type": "buttons",
                "buttons": ["홈으로"]
            }
        }
        
    return jsonify(response)

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=int(sys.argv[1]), debug=True)
